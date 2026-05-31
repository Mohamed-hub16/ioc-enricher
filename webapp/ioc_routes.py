import ipaddress
import logging
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, send_file, jsonify
from flask_login import login_required, current_user
from sqlalchemy import func as sqlfunc

logger = logging.getLogger(__name__)

from webapp import db
from webapp.models import IOCRecord, Comment
from src.parsers.ioc_parser import parse_iocs
from src.enrichers import ENRICHERS_BY_TYPE
from src.synthesis import groq_synthesizer

ioc_bp = Blueprint("ioc", __name__)

_MAX_FILE_BYTES = 512 * 1024   # 512 Ko
_MAX_BATCH      = 30


def _parse_upload(file_storage) -> tuple[list[str], str | None]:
    """Validate and parse an uploaded .txt file. Returns (values, error_message)."""
    filename = file_storage.filename or ""

    # 1. Extension — whitelist strict
    if not filename.lower().endswith(".txt"):
        return [], "Seuls les fichiers .txt sont acceptés."

    # 2. Read at most MAX+1 bytes to detect oversized files without loading everything
    raw_bytes = file_storage.read(_MAX_FILE_BYTES + 1)
    if len(raw_bytes) > _MAX_FILE_BYTES:
        return [], "Fichier trop volumineux (maximum 512 Ko)."

    # 3. Decode as UTF-8 — reject binary / unexpected encodings
    try:
        text = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return [], "Fichier illisible — encodage non supporté (UTF-8 attendu)."

    # 4. Parse lines — strip inline comments and blanks
    values = []
    for line in text.splitlines():
        line = line.split("#")[0].strip()
        if line:
            values.append(line)

    if not values:
        return [], "Fichier vide (aucun IOC trouvé — les lignes débutant par # sont ignorées)."

    if len(values) > _MAX_BATCH:
        return [], (
            f"Trop d'IOCs dans le fichier ({len(values)}) — "
            f"maximum {_MAX_BATCH} par import."
        )

    return values, None


def _parse_text_input(raw: str) -> list[str]:
    """Split a free-text input on commas and/or newlines, deduplicate, preserve order."""
    seen: set[str] = set()
    result: list[str] = []
    for part in re.split(r"[,\n\r]+", raw):
        part = part.strip()
        if part and part not in seen:
            seen.add(part)
            result.append(part)
    return result


def _fmt_ts(v) -> str:
    if not v:
        return "—"
    try:
        return datetime.utcfromtimestamp(int(v)).strftime("%Y-%m-%d")
    except Exception:
        return str(v)[:10]


def _build_summary(results: list[dict], ioc_type: str, ioc_value: str) -> dict | None:
    """Build the quick-summary table data for the IOC detail page."""
    src = {
        r["source"]: r.get("data") or {}
        for r in results
        if not r.get("error") and r.get("data")
    }
    vt  = src.get("VirusTotal", {})
    mal = vt.get("malicious", 0) or 0
    tot = vt.get("total_engines", 0) or 0
    vt_str = f"{mal}/{tot}" if tot else "—"

    if ioc_type == "ip":
        abuse = src.get("AbuseIPDB", {})
        ipapi = src.get("ip-api.com", {})
        asn_num = vt.get("asn") or ipapi.get("asn") or ""
        abuse_sc = abuse.get("abuse_confidence_score")
        fields = [
            ("IP",              ioc_value),
            ("Score AbuseIPDB", f"{abuse_sc}%" if abuse_sc is not None else "—"),
            ("Score VT",        vt_str),
            ("ISP",             ipapi.get("isp") or abuse.get("isp") or "—"),
            ("Usage Type",      ipapi.get("network_type") or "—"),
            ("ASN",             f"AS{asn_num}" if asn_num else "—"),
            ("AS Owner",        vt.get("as_owner") or ipapi.get("as_name") or "—"),
            ("Réseau",          vt.get("network") or "—"),
            ("Reverse DNS",     ipapi.get("reverse_dns") or "—"),
            ("Country",         ipapi.get("country") or abuse.get("country_name") or "—"),
            ("City",            ipapi.get("city") or "—"),
        ]

    elif ioc_type == "domain":
        cats = vt.get("categories") or {}
        cats_str = " · ".join(sorted({str(v) for v in cats.values()}))[:80] if cats else "—"
        rep = vt.get("reputation")
        fields = [
            ("Domaine",       ioc_value),
            ("Score VT",      vt_str),
            ("Réputation VT", str(rep) if rep is not None else "—"),
            ("Tags",          " · ".join(vt.get("tags") or []) or "—"),
            ("Catégories",    cats_str),
            ("Registrar",     vt.get("registrar") or "—"),
            ("Création",      _fmt_ts(vt.get("creation_date"))),
        ]

    elif ioc_type == "hash":
        exif = vt.get("exiftool") or {}
        sig  = vt.get("signature_info") or {}
        sub  = sig.get("subject") or ""
        cn   = (sub.split("CN=")[-1].split(",")[0] if "CN=" in sub
                else (sig.get("signers") or "").split(";")[0]).strip()
        signer_str = cn if cn else ("Non signé" if vt else "—")

        size = vt.get("file_size")
        size_str = f"{size:,}".replace(",", " ") + " o" if size else "—"

        sha256 = vt.get("sha256") or "—"
        sha256_display = sha256[:22] + "…" if len(sha256) > 22 else sha256

        fields = [
            ("MD5",             vt.get("md5") or ioc_value),
            ("SHA1",            vt.get("sha1") or "—"),
            ("SHA256",          sha256_display),
            ("Type",            vt.get("type_description") or vt.get("magic") or "—"),
            ("Taille",          size_str),
            ("Score VT",        vt_str),
            ("Famille",         vt.get("popular_threat_label") or "—"),
            ("Compilé le",      str(vt.get("pe_compilation_date") or "—")[:10]),
            ("1ère soumission", vt.get("first_submission_date") or "—"),
            ("Signé par",       signer_str),
            ("Company (Exif)",  exif.get("CompanyName") or "—"),
            ("Product (Exif)",  exif.get("ProductName") or "—"),
        ]
    else:
        return None

    headers = [f[0] for f in fields]
    values  = [str(f[1]) for f in fields]
    md_header = "| " + " | ".join(f"**{h}**" for h in headers) + " |"
    md_sep    = "| " + " | ".join(["---"] * len(headers)) + " |"
    md_row    = "| " + " | ".join(values) + " |"
    markdown  = "\n".join([md_header, md_sep, md_row])

    return {"fields": fields, "markdown": markdown}


def _is_private_ioc(value: str, ioc_type: str) -> bool:
    if ioc_type == "ip":
        try:
            ip = ipaddress.ip_address(value)
            return ip.is_private or ip.is_loopback or ip.is_reserved or ip.is_link_local
        except ValueError:
            return False
    if ioc_type == "domain":
        lower = value.lower()
        return lower == "localhost" or lower.endswith(".local") or lower.endswith(".internal")
    return False


def _compute_threat_score(raw_results: list[dict]) -> int:
    """0-100 threat score: AbuseIPDB + VirusTotal + MXToolBox blacklist bonus."""
    abuse_score = None
    vt_score = None
    bl_bonus = 0

    for res in raw_results:
        if res.get("error"):
            continue
        data = res.get("data", {})
        if res["source"] == "AbuseIPDB":
            abuse_score = data.get("abuse_confidence_score", 0) or 0
        elif res["source"] == "VirusTotal":
            malicious = data.get("malicious", 0) or 0
            total = data.get("total_engines", 0) or 0
            vt_score = round(malicious / total * 100) if total > 0 else 0
        elif res["source"] == "MXToolBox":
            bl_count = len(data.get("blacklisted_on") or [])
            bl_bonus = min(20, bl_count * 4)

    if abuse_score is not None and vt_score is not None:
        base = round(abuse_score * 0.5 + vt_score * 0.5)
    elif abuse_score is not None:
        base = int(abuse_score)
    elif vt_score is not None:
        base = int(vt_score)
    else:
        base = 0

    return min(100, base + bl_bonus)


def _all_sources_failed(raw_results: list[dict]) -> bool:
    return bool(raw_results) and all(res.get("error") for res in raw_results)


def _compute_score_breakdown(raw_results: list[dict]) -> dict:
    """Return per-source scores that contributed to the final threat score."""
    breakdown = {}
    for res in raw_results:
        if res.get("error") or not res.get("data"):
            continue
        data = res["data"]
        src = res["source"]
        if src == "AbuseIPDB":
            val = data.get("abuse_confidence_score")
            if val is not None:
                breakdown["AbuseIPDB"] = int(val)
        elif src == "VirusTotal":
            malicious = data.get("malicious", 0) or 0
            total = data.get("total_engines", 0) or 0
            if total > 0:
                breakdown["VirusTotal"] = round(malicious / total * 100)
        elif src == "URLhaus":
            if data.get("found"):
                breakdown["URLhaus"] = data.get("url_count", 0)
        elif src == "ThreatFox":
            if data.get("found"):
                breakdown["ThreatFox"] = len(data.get("results") or [])
        elif src == "GreyNoise":
            cls = data.get("classification")
            if cls == "malicious":
                breakdown["GreyNoise"] = "malicious"
            elif cls == "benign":
                breakdown["GreyNoise"] = "benign"
        elif src == "Shodan InternetDB":
            vulns = data.get("vulns") or []
            if vulns:
                breakdown["Shodan InternetDB"] = len(vulns)
    return breakdown


def _enrich_ioc(value: str, keys: dict | None = None, generate_ai: bool = True,
                level: str = "standard") -> tuple[str, list[dict], str | None, int, dict, str | None]:
    """Return (ioc_type, raw, paragraph, threat_score, score_breakdown, malware_family)."""
    iocs, _ = parse_iocs([value])
    if not iocs:
        raise ValueError(f"IOC non reconnu : {value!r}")

    ioc = iocs[0]
    enrichers = ENRICHERS_BY_TYPE.get(ioc.type, [])
    if not enrichers:
        raise ValueError(f"Aucun enrichisseur disponible pour le type '{ioc.type}'")

    # Run enrichers in parallel, preserving original ordering
    result_slots: list[object | None] = [None] * len(enrichers)

    def _run(idx, fn):
        return idx, fn(ioc, keys=keys)

    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = {pool.submit(_run, i, fn): i for i, fn in enumerate(enrichers)}
        for future in as_completed(futures):
            try:
                idx, r = future.result()
                result_slots[idx] = r
            except Exception as exc:
                logger.error("Enricher error: %s", exc)

    result_objects = [r for r in result_slots if r is not None]

    raw = [{"source": r.source, "data": r.data, "error": r.error} for r in result_objects]
    threat_score = _compute_threat_score(raw)
    score_breakdown = _compute_score_breakdown(raw)

    groq_key = keys.get("GROQ_API_KEY", "") if keys is not None else os.getenv("GROQ_API_KEY", "")
    if generate_ai and groq_key:
        synthesis = groq_synthesizer.synthesize_full(result_objects, threat_score, groq_key=groq_key, level=level)
        paragraph = synthesis["paragraph"]
        malware_family = synthesis["malware_family"]
    else:
        paragraph = None
        malware_family = None

    return ioc.type, raw, paragraph, threat_score, score_breakdown, malware_family


def _purge_stale_iocs() -> int:
    from webapp.models import STALE_DAYS
    cutoff = datetime.utcnow() - timedelta(days=STALE_DAYS)
    stale = IOCRecord.query.filter(IOCRecord.enriched_at < cutoff).all()
    count = len(stale)
    for record in stale:
        db.session.delete(record)
    if count:
        db.session.commit()
        logger.info("Purged %d stale IOC(s) older than %d days", count, STALE_DAYS)
    return count


@ioc_bp.route("/")
def index():
    _purge_stale_iocs()
    is_visitor    = not current_user.is_authenticated
    q             = request.args.get("q", "").strip()
    filt          = request.args.get("filter", "all")   # all | malicious | suspect | legitimate
    sort          = request.args.get("sort", "score")   # score | date | views
    type_filter   = request.args.get("type", "all")     # all | ip | domain | hash
    family_filter = request.args.get("family", "").strip()
    tag_filter    = request.args.get("tag", "").strip()

    query = IOCRecord.query
    if is_visitor:
        query = query.filter(IOCRecord.tags_json.ilike('%"demo"%'))
    if q:
        query = query.filter(IOCRecord.value.ilike(f"%{q}%"))
    if filt == "malicious":
        query = query.filter(IOCRecord.threat_score > 30)
    elif filt == "suspect":
        query = query.filter(IOCRecord.threat_score > 0, IOCRecord.threat_score <= 30)
    elif filt == "legitimate":
        query = query.filter(IOCRecord.threat_score == 0)
    if type_filter in ("ip", "domain", "hash"):
        query = query.filter(IOCRecord.ioc_type == type_filter)
    if family_filter:
        query = query.filter(IOCRecord.malware_family.ilike(f"%{family_filter}%"))
    if tag_filter:
        query = query.filter(IOCRecord.tags_json.ilike(f'%"{tag_filter}"%'))

    if sort == "score":
        query = query.order_by(IOCRecord.threat_score.desc())
    elif sort == "views":
        query = query.order_by(IOCRecord.view_count.desc())
    else:
        query = query.order_by(IOCRecord.enriched_at.desc())

    records = query.limit(100).all()

    _cnt_base = IOCRecord.query.filter(IOCRecord.tags_json.ilike('%"demo"%')) \
        if is_visitor else IOCRecord.query
    counts = {
        "all":       _cnt_base.count(),
        "malicious": _cnt_base.filter(IOCRecord.threat_score > 30).count(),
        "suspect":   _cnt_base.filter(
                         IOCRecord.threat_score > 0, IOCRecord.threat_score <= 30).count(),
        "legitimate":_cnt_base.filter(IOCRecord.threat_score == 0).count(),
        "ip":        _cnt_base.filter(IOCRecord.ioc_type == "ip").count(),
        "domain":    _cnt_base.filter(IOCRecord.ioc_type == "domain").count(),
        "hash":      _cnt_base.filter(IOCRecord.ioc_type == "hash").count(),
    }

    # Top families
    family_rows = (
        db.session.query(IOCRecord.malware_family,
                         sqlfunc.count(IOCRecord.id).label("n"))
        .filter(IOCRecord.malware_family.isnot(None), IOCRecord.malware_family != "")
        .group_by(IOCRecord.malware_family)
        .order_by(sqlfunc.count(IOCRecord.id).desc())
        .all()
    )
    families_top         = family_rows[:4]
    families_other_count = max(0, len(family_rows) - 4)

    # Tag frequency (aggregated from JSON blobs)
    tag_freq: dict[str, int] = {}
    for r in IOCRecord.query.filter(IOCRecord.tags_json.isnot(None)).all():
        for tag in r.get_tags():
            tag_freq[tag] = tag_freq.get(tag, 0) + 1
    sorted_tags = sorted(tag_freq.items(), key=lambda x: -x[1])
    tags_shown  = sorted_tags[:7]
    tags_extra  = max(0, len(sorted_tags) - 7)

    return render_template(
        "index.html",
        records=records, q=q, filt=filt, sort=sort,
        counts=counts, type_filter=type_filter,
        family_filter=family_filter, tag_filter=tag_filter,
        families_top=families_top, families_other_count=families_other_count,
        tags_shown=tags_shown, tags_extra=tags_extra,
        is_visitor=is_visitor,
    )


def _user_groq_key() -> str:
    """Resolve the Groq key for the current user (admin → .env, analyst → own)."""
    if current_user.is_admin:
        return os.getenv("GROQ_API_KEY", "")
    return current_user.get_api_keys().get("GROQ_API_KEY", "")


@ioc_bp.route("/ioc/<path:value>")
def detail(value: str):
    is_visitor = not current_user.is_authenticated

    if current_user.is_authenticated and not current_user.is_approved:
        abort(403)

    record = IOCRecord.query.filter_by(value=value).first_or_404()

    if is_visitor and "demo" not in record.get_tags():
        flash("Connectez-vous pour accéder à cet IOC.", "warning")
        return redirect(url_for("auth.login", next=request.path))

    record.view_count = (record.view_count or 0) + 1
    db.session.commit()
    results = record.get_results()
    summary = _build_summary(results, record.ioc_type, record.value)

    paragraphs = record.get_paragraphs()
    active_level = request.args.get("level", "")
    if active_level not in paragraphs:
        active_level = "standard" if "standard" in paragraphs else next(iter(paragraphs), "")

    if is_visitor:
        has_groq = False
    elif current_user.is_admin:
        has_groq = bool(os.getenv("GROQ_API_KEY"))
    else:
        has_groq = bool(current_user.groq_key_enc)

    return render_template(
        "ioc_detail.html", record=record, results=results, summary=summary,
        paragraphs=paragraphs, active_level=active_level, has_groq=has_groq,
        is_visitor=is_visitor,
    )


@ioc_bp.route("/ioc/<path:value>/analyze", methods=["POST"])
@login_required
def generate_analysis(value: str):
    """Generate one analysis level from already-stored raw_results — Groq only, no enricher calls."""
    if not current_user.is_approved:
        abort(403)
    record = IOCRecord.query.filter_by(value=value).first_or_404()

    level = request.form.get("level", "standard")
    if level not in ("brief", "standard", "detailed"):
        flash("Niveau d'analyse invalide.", "danger")
        return redirect(url_for("ioc.detail", value=value))

    groq_key = _user_groq_key()
    if not groq_key:
        flash("Aucune clé Groq configurée — impossible de générer l'analyse IA.", "warning")
        return redirect(url_for("ioc.detail", value=value))

    raw = record.get_results()
    if not raw:
        flash("Aucune donnée d'enrichissement à analyser.", "danger")
        return redirect(url_for("ioc.detail", value=value))

    from src.models import IOC, EnrichmentResult
    ioc_obj = IOC(value=record.value, type=record.ioc_type)
    result_objects = [
        EnrichmentResult(source=r["source"], ioc=ioc_obj,
                         data=r.get("data") or {}, error=r.get("error"))
        for r in raw
    ]

    paragraph = groq_synthesizer.synthesize(
        result_objects, threat_score=record.threat_score or 0,
        groq_key=groq_key, level=level,
    )
    if not paragraph:
        flash("La génération de l'analyse a échoué — réessayez dans un instant.", "danger")
        return redirect(url_for("ioc.detail", value=value))

    record.set_paragraph_level(level, paragraph)
    db.session.commit()
    _LEVEL_LABELS = {"brief": "Bref", "standard": "Standard", "detailed": "Détaillé"}
    flash(f"Analyse « {_LEVEL_LABELS[level]} » générée.", "success")
    return redirect(url_for("ioc.detail", value=value, level=level))


@ioc_bp.route("/enrich", methods=["GET", "POST"])
@login_required
def enrich():
    if not current_user.is_approved:
        flash("Votre compte n'est pas encore approuvé par un administrateur.", "warning")
        return redirect(url_for("ioc.index"))

    if request.method == "POST":
        # Analysts must have configured their own API keys
        if not current_user.is_admin and not current_user.has_api_keys:
            flash("Vous devez d'abord configurer vos clés API.", "warning")
            return redirect(url_for("auth.api_keys"))

        force       = request.form.get("force") == "1"
        generate_ai = request.form.get("generate_ai") == "1"
        level       = request.form.get("analysis_level", "standard")
        if level not in ("brief", "standard", "detailed"):
            level = "standard"
        keys        = current_user.get_api_keys() if not current_user.is_admin else None

        import json as _json
        try:
            default_tags = _json.loads(request.form.get("default_tags_json", "[]"))
            if not isinstance(default_tags, list):
                default_tags = []
        except Exception:
            default_tags = []

        # ── Collect IOC values: file upload takes priority over text input ──
        uploaded = request.files.get("ioc_file")
        if uploaded and uploaded.filename:
            values, file_error = _parse_upload(uploaded)
            if file_error:
                flash(file_error, "danger")
                return render_template("enrich.html", prefill="", max_batch=_MAX_BATCH)
        else:
            raw = request.form.get("ioc", "").strip()
            if not raw:
                flash("Veuillez saisir au moins un IOC.", "danger")
                return render_template("enrich.html", prefill="", max_batch=_MAX_BATCH)
            values = _parse_text_input(raw)

        if not values:
            flash("Aucun IOC valide trouvé.", "danger")
            return render_template("enrich.html", prefill="", max_batch=_MAX_BATCH)

        # ── Single IOC: original flow (redirect to detail page) ──
        if len(values) == 1:
            value = values[0]

            iocs, _ = parse_iocs([value])
            if not iocs:
                flash(f"IOC non reconnu : {value!r}", "danger")
                return render_template("enrich.html", prefill=value, max_batch=_MAX_BATCH)

            ioc_type = iocs[0].type

            if _is_private_ioc(value, ioc_type):
                return render_template(
                    "enrich.html",
                    prefill=value, max_batch=_MAX_BATCH,
                    no_info=True,
                    no_info_reason="Adresse privée ou locale (RFC1918 / loopback / .local)",
                )

            existing = IOCRecord.query.filter_by(value=value).first()
            if existing and not force:
                flash(
                    f"IOC déjà en base (enrichi il y a {existing.age_days} jour(s)). "
                    "Résultat affiché depuis le cache.",
                    "info",
                )
                return redirect(url_for("ioc.detail", value=value))

            try:
                ioc_type, raw, paragraph, threat_score, score_breakdown, malware_family = _enrich_ioc(value, keys=keys, generate_ai=generate_ai, level=level)
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template("enrich.html", prefill=value, max_batch=_MAX_BATCH)
            except Exception as exc:
                logger.error("Enrichment error for %r: %s", value, exc)
                flash("Une erreur est survenue lors de l'enrichissement.", "danger")
                return render_template("enrich.html", prefill=value, max_batch=_MAX_BATCH)

            if _all_sources_failed(raw):
                return render_template(
                    "enrich.html",
                    prefill=value, max_batch=_MAX_BATCH,
                    no_info=True,
                    no_info_reason="Aucune source n'a retourné d'information pour cet IOC.",
                )

            existing = IOCRecord.query.filter_by(value=value).first()
            if existing:
                existing.enriched_at = datetime.utcnow()
                existing.enriched_by = current_user.email
                existing.set_results(raw)
                existing.set_paragraph_level(level, paragraph)
                existing.threat_score = threat_score
                existing.malware_family = malware_family
                existing.verdict = "malicious" if threat_score > 0 else "clean"
                existing.set_score_breakdown(score_breakdown)
            else:
                record = IOCRecord(
                    value=value, ioc_type=ioc_type,
                    enriched_by=current_user.email,
                    threat_score=threat_score,
                    malware_family=malware_family,
                    verdict="malicious" if threat_score > 0 else "clean",
                )
                record.set_results(raw)
                record.set_paragraph_level(level, paragraph)
                record.set_score_breakdown(score_breakdown)
                if default_tags:
                    record.set_tags(default_tags)
                db.session.add(record)

            db.session.commit()
            flash("Enrichissement terminé avec succès.", "success")
            return redirect(url_for("ioc.detail", value=value))

        # ── Bulk: process all IOCs and display a results table ──
        bulk_results = []
        for value in values[:_MAX_BATCH]:
            entry = {"value": value, "ioc_type": "?", "threat_score": None,
                     "status": "error", "message": ""}

            iocs, _ = parse_iocs([value])
            if not iocs:
                entry["status"] = "skipped"
                entry["message"] = "IOC non reconnu"
                bulk_results.append(entry)
                continue

            ioc_type = iocs[0].type
            entry["ioc_type"] = ioc_type

            if _is_private_ioc(value, ioc_type):
                entry["status"] = "skipped"
                entry["message"] = "Adresse privée / locale"
                bulk_results.append(entry)
                continue

            existing = IOCRecord.query.filter_by(value=value).first()
            if existing and not force:
                entry["status"] = "cached"
                entry["threat_score"] = existing.threat_score
                bulk_results.append(entry)
                continue

            try:
                ioc_type, raw_res, paragraph, threat_score, score_breakdown, malware_family = _enrich_ioc(value, keys=keys, generate_ai=generate_ai, level=level)
            except Exception as exc:
                logger.error("Bulk enrichment error for %r: %s", value, exc)
                entry["message"] = "Erreur d'enrichissement"
                bulk_results.append(entry)
                continue

            if _all_sources_failed(raw_res):
                entry["status"] = "skipped"
                entry["message"] = "Aucune source n'a répondu"
                bulk_results.append(entry)
                continue

            now = datetime.utcnow()
            if existing:
                existing.enriched_at = now
                existing.enriched_by = current_user.email
                existing.set_results(raw_res)
                existing.set_paragraph_level(level, paragraph)
                existing.threat_score = threat_score
                existing.malware_family = malware_family
                existing.verdict = "malicious" if threat_score > 0 else "clean"
                existing.set_score_breakdown(score_breakdown)
                entry["status"] = "updated"
            else:
                record = IOCRecord(
                    value=value, ioc_type=ioc_type,
                    enriched_by=current_user.email,
                    threat_score=threat_score,
                    malware_family=malware_family,
                    verdict="malicious" if threat_score > 0 else "clean",
                )
                record.set_results(raw_res)
                record.set_paragraph_level(level, paragraph)
                record.set_score_breakdown(score_breakdown)
                if default_tags:
                    record.set_tags(default_tags)
                db.session.add(record)
                entry["status"] = "new"

            entry["threat_score"] = threat_score
            bulk_results.append(entry)

        db.session.commit()
        return render_template("enrich.html", bulk_results=bulk_results, max_batch=_MAX_BATCH)

    keys = current_user.get_api_keys() if not current_user.is_admin else None
    has_groq = bool(
        keys.get("GROQ_API_KEY", "") if keys is not None else os.getenv("GROQ_API_KEY")
    )
    return render_template("enrich.html", prefill=request.args.get("ioc", ""),
                           max_batch=_MAX_BATCH, has_groq=has_groq)


@ioc_bp.route("/ioc/<path:value>/delete", methods=["POST"])
@login_required
def delete_ioc(value: str):
    if not current_user.is_admin:
        abort(403)
    record = IOCRecord.query.filter_by(value=value).first_or_404()
    Comment.query.filter_by(ioc_record_id=record.id).delete()
    db.session.delete(record)
    db.session.commit()
    flash(f"IOC « {value} » supprimé.", "success")
    return redirect(url_for("ioc.index"))


@ioc_bp.route("/export/excel")
@login_required
def export_excel():
    if not current_user.is_approved:
        abort(403)
    family_filter = request.args.get("family", "").strip()

    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    IP_HEADERS = ["Valeur", "ISP", "Score global", "Score AbuseIPDB", "Score VirusTotal",
                  "Vues", "Malveillant", "Enrichi le", "Enrichi par"]
    OTHER_HEADERS = ["Valeur", "Score global", "Score AbuseIPDB", "Score VirusTotal",
                     "Vues", "Malveillant", "Enrichi le", "Enrichi par"]
    IP_WIDTHS    = [38, 30, 13, 16, 16, 8, 12, 20, 28]
    OTHER_WIDTHS = [45,     13, 16, 16, 8, 12, 20, 28]

    HDR_FONT = Font(bold=True, color="C9D1D9")
    HDR_FILL = PatternFill("solid", fgColor="1C2128")
    HDR_ALIGN = Alignment(horizontal="center")
    RED_FONT  = Font(color="DA3633", bold=True)

    def _extract_row_data(raw_results: list[dict], ioc_type: str) -> tuple[str, str, str, str]:
        """Return (isp, abuse_score, vt_score) extracted from raw results."""
        isp = ""
        abuse_score = ""
        vt_score = ""
        for res in raw_results:
            if res.get("error") or not res.get("data"):
                continue
            d = res["data"]
            if res["source"] == "ip-api.com":
                isp = d.get("isp") or ""
            elif res["source"] == "AbuseIPDB":
                val = d.get("abuse_confidence_score")
                if val is not None:
                    abuse_score = str(val)
                if not isp:
                    isp = d.get("isp") or ""
            elif res["source"] == "VirusTotal":
                mal = d.get("malicious", 0) or 0
                tot = d.get("total_engines", 0) or 0
                if tot > 0:
                    vt_score = str(round(mal / tot * 100))
                elif mal == 0:
                    vt_score = "0"
        return isp, abuse_score, vt_score

    for ioc_type in ("ip", "domain", "hash"):
        is_ip = ioc_type == "ip"
        headers   = IP_HEADERS if is_ip else OTHER_HEADERS
        col_widths = IP_WIDTHS  if is_ip else OTHER_WIDTHS

        ws = wb.create_sheet(title=ioc_type.upper() + "s")
        ws.append(headers)
        for i, cell in enumerate(ws[1], start=1):
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN
            ws.column_dimensions[cell.column_letter].width = col_widths[i - 1]

        rq = IOCRecord.query.filter_by(ioc_type=ioc_type)
        if family_filter:
            rq = rq.filter(IOCRecord.malware_family.ilike(f"%{family_filter}%"))
        records = rq.order_by(IOCRecord.threat_score.desc(), IOCRecord.enriched_at.desc()).all()
        for r in records:
            isp, abuse_score, vt_score = _extract_row_data(r.get_results(), ioc_type)
            if is_ip:
                row = [r.value, isp, r.threat_score or 0, abuse_score, vt_score,
                       r.view_count or 0, "Oui" if r.is_malicious else "Non",
                       r.enriched_at.strftime("%Y-%m-%d %H:%M UTC"), r.enriched_by or ""]
                mal_cols = (1, 3, 7)   # Valeur, Score global, Malveillant
            else:
                row = [r.value, r.threat_score or 0, abuse_score, vt_score,
                       r.view_count or 0, "Oui" if r.is_malicious else "Non",
                       r.enriched_at.strftime("%Y-%m-%d %H:%M UTC"), r.enriched_by or ""]
                mal_cols = (1, 2, 6)
            ws.append(row)
            if r.is_malicious:
                for col in mal_cols:
                    ws.cell(row=ws.max_row, column=col).font = RED_FONT

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ioc_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@ioc_bp.route("/export/markdown")
@login_required
def export_markdown():
    if not current_user.is_approved:
        abort(403)

    from io import BytesIO

    records = (
        IOCRecord.query
        .order_by(IOCRecord.threat_score.desc(), IOCRecord.enriched_at.desc())
        .all()
    )

    lines = [
        "# IOC Export",
        "",
        f"Exporté le {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC — {len(records)} IOCs",
        "",
        "| Valeur | Type | Score | Verdict | Famille | Enrichi le |",
        "|--------|------|------:|---------|---------|------------|",
    ]
    for r in records:
        if (r.threat_score or 0) > 30:
            verdict = "Malveillant"
        elif (r.threat_score or 0) > 0:
            verdict = "Suspect"
        else:
            verdict = "Légitime"
        lines.append(
            f"| `{r.value}` | {r.ioc_type} | {r.threat_score or 0} | {verdict} "
            f"| {r.malware_family or '—'} "
            f"| {r.enriched_at.strftime('%Y-%m-%d')} |"
        )

    content = "\n".join(lines) + "\n"
    filename = f"ioc_export_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.md"
    return send_file(
        BytesIO(content.encode("utf-8")),
        download_name=filename,
        as_attachment=True,
        mimetype="text/markdown; charset=utf-8",
    )


@ioc_bp.route("/ioc/<path:value>/tags", methods=["POST"])
@login_required
def set_tags(value: str):
    if not current_user.is_approved:
        abort(403)
    record = IOCRecord.query.filter_by(value=value).first_or_404()
    raw_tags = request.form.get("tags", "")
    tags = [t.strip()[:64] for t in re.split(r"[,;]+", raw_tags) if t.strip()][:20]
    record.set_tags(tags)
    db.session.commit()
    flash("Tags mis à jour.", "success")
    return redirect(url_for("ioc.detail", value=value))


def _rel_time(dt: datetime) -> str:
    s = int((datetime.utcnow() - dt).total_seconds())
    if s < 60:   return "à l'instant"
    if s < 3600: return f"il y a {s // 60}min"
    if s < 86400:return f"il y a {s // 3600}h"
    d = s // 86400
    if d < 30:   return f"il y a {d}j"
    if d < 365:  return f"il y a {d // 30} mois"
    return f"il y a {d // 365} an{'s' if d // 365 > 1 else ''}"


@ioc_bp.route("/pivot/families")
@login_required
def pivot_families():
    if not current_user.is_approved:
        abort(403)
    rows = (
        db.session.query(
            IOCRecord.malware_family,
            sqlfunc.count(IOCRecord.id).label("total"),
            sqlfunc.sum(sqlfunc.cast(IOCRecord.threat_score > 30, db.Integer)).label("malicious"),
            sqlfunc.max(IOCRecord.enriched_at).label("last_seen"),
        )
        .filter(IOCRecord.malware_family.isnot(None), IOCRecord.malware_family != "")
        .group_by(IOCRecord.malware_family)
        .order_by(sqlfunc.count(IOCRecord.id).desc())
        .all()
    )
    return render_template("pivot_families.html", families=rows,
                           now_utc=datetime.utcnow())


@ioc_bp.route("/pivot/family/<path:family>")
@login_required
def pivot_family(family: str):
    if not current_user.is_approved:
        abort(403)
    sort = request.args.get("sort", "score")

    q = IOCRecord.query.filter(IOCRecord.malware_family.ilike(f"%{family}%"))
    if sort == "date":
        q = q.order_by(IOCRecord.enriched_at.desc())
    elif sort == "type":
        q = q.order_by(IOCRecord.ioc_type, IOCRecord.threat_score.desc())
    else:
        q = q.order_by(IOCRecord.threat_score.desc(), IOCRecord.enriched_at.desc())

    records = q.limit(200).all()

    stats = {}
    synthesis = None

    if records:
        stats = {
            "total":     len(records),
            "ip":        sum(1 for r in records if r.ioc_type == "ip"),
            "domain":    sum(1 for r in records if r.ioc_type == "domain"),
            "hash":      sum(1 for r in records if r.ioc_type == "hash"),
            "malicious": sum(1 for r in records if (r.threat_score or 0) > 30),
            "first_seen":     min(r.enriched_at for r in records),
            "last_seen":      max(r.enriched_at for r in records),
            "last_seen_rel":  _rel_time(max(r.enriched_at for r in records)),
            "first_seen_fmt": min(r.enriched_at for r in records).strftime("%d %b %Y"),
        }

        groq_key = os.getenv("GROQ_API_KEY", "")
        if groq_key:
            ctx_lines = [
                f"{stats['total']} IOCs "
                f"({stats['ip']} IPs, {stats['domain']} domaines, {stats['hash']} hashes)",
                f"Malveillants (score > 30) : {stats['malicious']}",
                "",
            ]
            for r in records[:10]:
                ctx_lines.append(
                    f"[{r.ioc_type.upper()} score={r.threat_score or 0}] {r.value}"
                )
                if r.paragraph:
                    ctx_lines.append(f"  → {r.paragraph[:400]}")
                ctx_lines.append("")
            synthesis = groq_synthesizer.synthesize_family(
                family, "\n".join(ctx_lines), groq_key=groq_key
            )

    return render_template("pivot_family.html", family=family, records=records,
                           stats=stats, sort=sort, synthesis=synthesis)


@ioc_bp.route("/ioc/<path:value>/comment", methods=["POST"])
@login_required
def add_comment(value: str):
    if not current_user.is_approved:
        abort(403)
    record = IOCRecord.query.filter_by(value=value).first_or_404()
    content = request.form.get("content", "").strip()
    if not content:
        flash("Le commentaire ne peut pas être vide.", "danger")
        return redirect(url_for("ioc.detail", value=value))
    comment = Comment(
        ioc_record_id=record.id,
        author=current_user.email,
        content=content,
        enriched_at_snapshot=record.enriched_at,
    )
    db.session.add(comment)
    db.session.commit()
    flash("Commentaire ajouté.", "success")
    return redirect(url_for("ioc.detail", value=value))
